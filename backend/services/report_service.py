import os
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

class ReportService:
    @staticmethod
    def ensure_directory(directory_path):
        if not os.path.exists(directory_path):
            os.makedirs(directory_path, exist_ok=True)

    @staticmethod
    def generate_csv(calc_data, filepath):
        """
        Generates a standard CSV report detailing the cost calculations and comparison table.
        """
        ReportService.ensure_directory(os.path.dirname(filepath))
        
        with open(filepath, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header Info
            writer.writerow(["CLOUD COST OPTIMIZER - CONFIGURATION REPORT"])
            writer.writerow(["Generated At", datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')])
            writer.writerow(["Selected Provider", calc_data['selected_provider_cost']['provider_name']])
            writer.writerow(["Selected Region", calc_data['selected_provider_cost']['breakdown'].get('region_code', 'Selected Region')])
            writer.writerow(["Monthly Cost", f"${calc_data['selected_provider_cost']['monthly_cost']}"])
            writer.writerow(["Annual Cost", f"${calc_data['selected_provider_cost']['annual_cost']}"])
            writer.writerow([])
            
            # Configuration Details
            writer.writerow(["RESOURCE SPECIFICATIONS"])
            writer.writerow(["Metric", "Monthly Cost"])
            breakdown = calc_data['selected_provider_cost']['breakdown']
            for metric, cost in breakdown.items():
                writer.writerow([metric.replace('_', ' ').capitalize(), f"${cost}"])
            writer.writerow([])
            
            # Comparison Table
            writer.writerow(["MULTI-CLOUD PRICING COMPARISON"])
            writer.writerow(["Rank", "Provider", "Compute Service", "Storage Service", "Monthly Cost", "Annual Cost", "Savings VS Selected", "Cheapest?"])
            for comp in calc_data['comparison']:
                writer.writerow([
                    comp['ranking'],
                    comp['provider_name'],
                    comp['compute_service'],
                    comp['storage_service'],
                    f"${comp['monthly_cost']}",
                    f"${comp['annual_cost']}",
                    f"${comp['estimated_savings']}",
                    "YES" if comp['cheapest_option'] else "NO"
                ])
            writer.writerow([])
            
            # Recommendations
            writer.writerow(["FINOPS ARCHITECTURAL RECOMMENDATIONS"])
            writer.writerow(["Type", "Title", "Complexity", "Impact", "Potential Savings (Monthly)", "Reasoning"])
            for rec in calc_data['recommendations']:
                writer.writerow([
                    rec['type'],
                    rec['title'],
                    rec['complexity'],
                    rec['impact'],
                    f"${rec['potential_savings']}",
                    rec['reasoning']
                ])
                
        return filepath

    @staticmethod
    def generate_excel(calc_data, filepath):
        """
        Generates a highly styled Excel sheet.
        """
        ReportService.ensure_directory(os.path.dirname(filepath))
        wb = Workbook()
        
        # Sheet 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Cost Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        # Colors & Fonts
        navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        gray_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        accent_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
        
        font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_header = Font(name="Calibri", size=12, bold=True, color="1E293B")
        font_tbl_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        # Title Block
        ws1.merge_cells("A1:D2")
        title_cell = ws1["A1"]
        title_cell.value = "Enterprise Cloud Cost Optimizer - Executive Report"
        title_cell.font = font_title
        title_cell.fill = navy_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws1["A4"] = "Metadata"
        ws1["A4"].font = font_header
        
        metadata = [
            ("Generated At:", datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')),
            ("Selected Provider:", calc_data['selected_provider_cost']['provider_name']),
            ("Baseline Monthly Spend:", calc_data['selected_provider_cost']['monthly_cost']),
            ("Baseline Annual Spend:", calc_data['selected_provider_cost']['annual_cost'])
        ]
        
        for r_idx, (k, v) in enumerate(metadata, start=5):
            ws1.cell(row=r_idx, column=1, value=k).font = font_bold
            ws1.cell(row=r_idx, column=2, value=v).font = font_regular
            if isinstance(v, (int, float)):
                ws1.cell(row=r_idx, column=2).number_format = '$#,##0.00'
                
        # Subtitle Resource Breakdown
        ws1["A10"] = "Selected Resource Breakdown"
        ws1["A10"].font = font_header
        
        breakdown_headers = ["Resource Category", "Monthly Baseline Cost"]
        for col, h in enumerate(breakdown_headers, start=1):
            cell = ws1.cell(row=11, column=col, value=h)
            cell.font = font_tbl_header
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center")
            
        breakdown_data = calc_data['selected_provider_cost']['breakdown']
        r_idx = 12
        for metric, cost in breakdown_data.items():
            ws1.cell(row=r_idx, column=1, value=metric.replace('_', ' ').capitalize()).font = font_regular
            c_cell = ws1.cell(row=r_idx, column=2, value=cost)
            c_cell.font = font_regular
            c_cell.number_format = '$#,##0.00'
            ws1.cell(row=r_idx, column=1).border = thin_border
            ws1.cell(row=r_idx, column=2).border = thin_border
            r_idx += 1
            
        # Sheet 2: Cloud Comparison Engine
        ws2 = wb.create_sheet(title="Provider Comparison")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2["A1"] = "Global Cloud Comparison Engine Matrix"
        ws2["A1"].font = Font(name="Calibri", size=14, bold=True)
        
        comp_headers = ["Rank", "Provider Name", "Compute Equivalent", "Storage Equivalent", "Monthly Spend", "Annual Spend", "Monthly Savings", "Price Diff %"]
        for col_idx, h in enumerate(comp_headers, start=1):
            cell = ws2.cell(row=3, column=col_idx, value=h)
            cell.font = font_tbl_header
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center")
            
        for r_idx, comp in enumerate(calc_data['comparison'], start=4):
            ws2.cell(row=r_idx, column=1, value=comp['ranking']).font = font_regular
            ws2.cell(row=r_idx, column=2, value=comp['provider_name']).font = font_bold
            ws2.cell(row=r_idx, column=3, value=comp['compute_service']).font = font_regular
            ws2.cell(row=r_idx, column=4, value=comp['storage_service']).font = font_regular
            
            c5 = ws2.cell(row=r_idx, column=5, value=comp['monthly_cost'])
            c5.number_format = '$#,##0.00'
            c5.font = font_regular
            
            c6 = ws2.cell(row=r_idx, column=6, value=comp['annual_cost'])
            c6.number_format = '$#,##0.00'
            c6.font = font_regular
            
            c7 = ws2.cell(row=r_idx, column=7, value=comp['estimated_savings'])
            c7.number_format = '$#,##0.00'
            c7.font = font_regular
            
            c8 = ws2.cell(row=r_idx, column=8, value=comp['diff_percent'] / 100.0)
            c8.number_format = '0.0%'
            c8.font = font_regular
            
            # Highlight Cheapest
            if comp['cheapest_option']:
                for col in range(1, 9):
                    ws2.cell(row=r_idx, column=col).fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                    
            for col in range(1, 9):
                ws2.cell(row=r_idx, column=col).border = thin_border
                
        # Sheet 3: Recommendations
        ws3 = wb.create_sheet(title="FinOps Savings Plan")
        ws3.views.sheetView[0].showGridLines = True
        
        ws3["A1"] = "Actionable FinOps Optimization Directives"
        ws3["A1"].font = Font(name="Calibri", size=14, bold=True)
        
        rec_headers = ["Domain", "Optimization Directive", "Potential Monthly Savings", "Implementation Complexity", "Impact Severity"]
        for col_idx, h in enumerate(rec_headers, start=1):
            cell = ws3.cell(row=3, column=col_idx, value=h)
            cell.font = font_tbl_header
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center")
            
        for r_idx, rec in enumerate(calc_data['recommendations'], start=4):
            ws3.cell(row=r_idx, column=1, value=rec['type']).font = font_bold
            ws3.cell(row=r_idx, column=2, value=rec['title']).font = font_regular
            
            sav_cell = ws3.cell(row=r_idx, column=3, value=rec['potential_savings'])
            sav_cell.number_format = '$#,##0.00'
            sav_cell.font = font_bold
            
            ws3.cell(row=r_idx, column=4, value=rec['complexity']).font = font_regular
            ws3.cell(row=r_idx, column=5, value=rec['impact']).font = font_regular
            
            for col in range(1, 6):
                ws3.cell(row=r_idx, column=col).border = thin_border
                ws3.cell(row=r_idx, column=col).fill = gray_fill
                
        # Autofit column widths
        from openpyxl.utils import get_column_letter
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        wb.save(filepath)
        return filepath

    @staticmethod
    def generate_pdf(calc_data, filepath):
        """
        Generates a professional PDF document.
        """
        ReportService.ensure_directory(os.path.dirname(filepath))
        
        doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        story = []
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            name='TitleStyle',
            fontName='Helvetica-Bold',
            fontSize=22,
            textColor=colors.HexColor('#1E293B'),
            spaceAfter=15,
            alignment=0 # Left
        )
        subtitle_style = ParagraphStyle(
            name='SubtitleStyle',
            fontName='Helvetica',
            fontSize=11,
            textColor=colors.HexColor('#64748B'),
            spaceAfter=25
        )
        h1_style = ParagraphStyle(
            name='Heading1Style',
            fontName='Helvetica-Bold',
            fontSize=15,
            textColor=colors.HexColor('#0F172A'),
            spaceBefore=15,
            spaceAfter=10
        )
        body_style = ParagraphStyle(
            name='BodyStyle',
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#334155'),
            spaceAfter=6
        )
        bold_body_style = ParagraphStyle(
            name='BoldBodyStyle',
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.HexColor('#0F172A'),
            spaceAfter=6
        )
        
        # Cover/Header
        story.append(Paragraph("Enterprise Cloud Cost Optimization Report", title_style))
        story.append(Paragraph(f"Generated on {datetime.utcnow().strftime('%B %d, %Y - %H:%M:%S UTC')} for enterprise infrastructure benchmarking.", subtitle_style))
        story.append(Spacer(1, 10))
        
        # Section 1: Selected Architecture Cost
        story.append(Paragraph("1. Active Configuration Summary", h1_style))
        story.append(Paragraph(f"<b>Baseline Provider:</b> {calc_data['selected_provider_cost']['provider_name']}", body_style))
        story.append(Paragraph(f"<b>Monthly Cost Estimate:</b> ${calc_data['selected_provider_cost']['monthly_cost']}", body_style))
        story.append(Paragraph(f"<b>Annualized Cost Projection:</b> ${calc_data['selected_provider_cost']['annual_cost']}", body_style))
        story.append(Spacer(1, 10))
        
        # Resource Table
        story.append(Paragraph("Selected Resource Allocation Breakdown:", bold_body_style))
        breakdown_data = [["Resource Category", "Monthly Estimated Cost (USD)"]]
        for k, v in calc_data['selected_provider_cost']['breakdown'].items():
            if isinstance(v, (int, float)):
                breakdown_data.append([k.replace('_', ' ').capitalize(), f"${v:.2f}"])
            else:
                breakdown_data.append([k.replace('_', ' ').capitalize(), str(v)])
            
        breakdown_table = Table(breakdown_data, colWidths=[200, 150])
        breakdown_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F8FAFC')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
        ]))
        story.append(breakdown_table)
        story.append(Spacer(1, 15))
        
        # Section 2: Multi-Cloud Benchmarking
        story.append(Paragraph("2. Multi-Cloud Pricing Comparison Matrix", h1_style))
        story.append(Paragraph("Calculated equivalent workload cost across global provider networks:", body_style))
        
        comp_data = [["Rank", "Provider", "Compute", "Storage", "Monthly", "Annual", "Savings"]]
        # Limit to top 12 providers in PDF to ensure readable layout
        for comp in calc_data['comparison'][:12]:
            comp_data.append([
                str(comp['ranking']),
                comp['provider_name'][:18],
                comp['compute_service'][:15],
                comp['storage_service'][:15],
                f"${comp['monthly_cost']:.2f}",
                f"${comp['annual_cost']:.2f}",
                f"${comp['estimated_savings']:.2f}"
            ])
            
        comp_table = Table(comp_data, colWidths=[30, 110, 90, 90, 65, 75, 65])
        comp_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            # Highlight first row (cheapest)
            ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#DCFCE7')),
            ('TEXTCOLOR', (0,1), (-1,1), colors.HexColor('#15803D')),
        ]))
        story.append(comp_table)
        story.append(Spacer(1, 15))
        
        # Section 3: Recommendations
        if calc_data['recommendations']:
            story.append(Paragraph("3. Actionable FinOps Recommendations", h1_style))
            story.append(Paragraph("Leverage these strategic architectural adjustments to secure additional cost optimization:", body_style))
            
            for rec in calc_data['recommendations']:
                rec_title = f"<b>[{rec['type']}] {rec['title']}</b>"
                story.append(Paragraph(rec_title, bold_body_style))
                story.append(Paragraph(f"Complexity: <b>{rec['complexity']}</b> | Impact: <b>{rec['impact']}</b> | Est. Savings: <b>${rec['potential_savings']:.2f}/mo</b>", body_style))
                story.append(Paragraph(rec['description'], body_style))
                story.append(Spacer(1, 8))
                
        # Build Document
        doc.build(story)
        return filepath
